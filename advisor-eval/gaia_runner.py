"""Agentic GAIA runner with tool-use and advisor escalation hooks."""

from __future__ import annotations

import ast
import csv
import html
import io
import json
import math
import mimetypes
import re
import time
import urllib.error
import urllib.parse
import urllib.request
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from huggingface_hub import hf_hub_download
from openai import OpenAI

from advisor import AdvisorAgent
from policies import EscalationPolicy

GAIA_SYSTEM_PROMPT = """\
You are an agentic problem solver. You can think, use tools, and provide a final answer.

At each turn, output ONLY a valid JSON object with this schema:
{
  "thought": "brief reasoning",
  "request_advisor": false,
  "action": "tool" | "final",
  "tool_name": "wiki_search" | "wiki_lookup" | "web_search" | "web_fetch_url" | "arxiv_search" | "arxiv_fetch" | "github_issue_fetch" | "read_attachment" | "calculator",
  "tool_input": "input for tool call",
  "final_answer": "answer string when action=final"
}

Tool usage guide:
- wiki_search: tool_input = plain search string (e.g. "Mercedes Sosa discography").
- wiki_lookup: tool_input = exact Wikipedia page title.
- web_search: tool_input = plain web search query string.
- web_fetch_url: tool_input = an https URL. Works for HTML, PDF, XLSX, CSV and plain text (content is auto-extracted).
- arxiv_search: tool_input = plain query OR arXiv fielded query (e.g. "cat:physics.soc-ph AND submittedDate:2016-08-11").
- arxiv_fetch: tool_input = arXiv id like "2401.12345".
- github_issue_fetch: use when a question mentions a specific GitHub issue/PR or needs repo metadata. tool_input = JSON {"repo":"owner/repo","issue_number":123} or the string "owner/repo#123".
- read_attachment: call ONLY when the task metadata lists an attachment_file_name. tool_input can be empty to read the default attachment, or JSON {"file_name":"...","sheet":"Sheet1","max_rows":200,"member":"inner.txt"}.
- calculator: tool_input = arithmetic expression (e.g. "23*17+4").

Rules:
- If action is "tool", provide tool_name and tool_input.
- If action is "final", provide final_answer and do NOT call a tool.
- For structured tool inputs, pass JSON in tool_input (string containing valid JSON).
- Do NOT call read_attachment when the task has no attachment_file_name.
- Never repeat the same (tool_name, tool_input) you already tried; reformulate or switch tools.
- Prefer this fallback order on web failures: web_search -> wiki_search/wiki_lookup -> arxiv_search (when academic).
- Avoid web_fetch_url on hosts that already returned 403/404/empty twice.
- For non-trivial factual questions, gather at least two corroborating signals before action=final.

Set request_advisor=true if ANY of these holds; otherwise keep it false:
- the same query returned no useful result twice;
- a tool returned an error twice;
- you are unsure how to format the final answer;
- you are about to emit a final answer with only one source or with hedging.

Final-answer format rubric (follow the question literally):
- If it asks for a number with no units, output digits only (e.g. "41", "0.1777"); obey any requested rounding or units like "thousand hours".
- If it asks for a comma-separated list, output "a, b, c" with a space after each comma and follow the requested ordering.
- If it asks for a name, output only the name with no title/honorific.
- If it asks for an exact phrase as it appears in a document, output only that phrase with no scene directive (no "INT.", "EXT.", or trailing "- DAY/NIGHT").
- Strip trailing units (m^3, kg, km, %) unless the question explicitly asks to include them.
- Before emitting action=final, silently check that your answer matches the requested format.\
"""


@dataclass
class GaiaRunResult:
    prediction: str | None = None
    total_exec_latency: float = 0.0
    total_adv_latency: float = 0.0
    total_exec_prompt: int = 0
    total_exec_completion: int = 0
    total_adv_prompt: int = 0
    total_adv_completion: int = 0
    advisor_calls: int = 0
    advisor_calls_after_error: int = 0
    advisor_guidance: list[dict[str, Any]] = field(default_factory=list)
    step_latencies: list[float] = field(default_factory=list)
    confidence_scores: list[float] = field(default_factory=list)
    tool_calls: int = 0
    tool_errors: int = 0
    dead_end_count: int = 0
    recovery_success: bool = False
    tool_trace: list[dict[str, Any]] = field(default_factory=list)
    repeated_query_violations: int = 0
    blocked_host_rehits: int = 0
    advisor_followed_first_step_count: int = 0
    advisor_first_step_total: int = 0


_CODE_FENCE_RE = re.compile(r"(?is)^```(?:json)?\s*(.*?)\s*```$")


def _strip_code_fence(text: str) -> str:
    text = text.strip()
    m = _CODE_FENCE_RE.match(text)
    if m:
        return m.group(1).strip()
    return text


def _find_first_json_object(text: str) -> str | None:
    """Return the first balanced {...} substring in text, respecting strings."""
    start = text.find("{")
    if start < 0:
        return None
    depth = 0
    in_str = False
    escape = False
    for i in range(start, len(text)):
        ch = text[i]
        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
        elif ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return text[start : i + 1]
    return None


def _extract_json(text: str) -> dict[str, Any] | None:
    """Parse a JSON object from model text, tolerating markdown fences and
    duplicate / concatenated objects (gpt-5.4-nano sometimes emits two)."""
    if not text:
        return None
    text = _strip_code_fence(text.strip())
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    # Try just the first balanced JSON object.
    snippet = _find_first_json_object(text)
    if snippet:
        try:
            parsed = json.loads(snippet)
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    # Last resort: slice from first '{' to last '}' (legacy fallback).
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        try:
            parsed = json.loads(text[start : end + 1])
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            return None
    return None


_HEDGE_PATTERNS = (
    "i can't",
    "i cannot",
    "i am unable",
    "i'm unable",
    "unable to",
    "i need",
    "i don't have",
    "cannot be determined",
    "no attachment",
    "not enough information",
)


def _is_hedged_answer(answer: str | None) -> bool:
    if answer is None:
        return True
    text = answer.strip().lower()
    if text in {"", "none", "unknown", "n/a", "null"}:
        return True
    return any(p in text for p in _HEDGE_PATTERNS)


def _short(s: str, limit: int = 120) -> str:
    s = (s or "").replace("\n", " ").strip()
    if len(s) <= limit:
        return s
    return s[: limit - 1] + "…"


def _normalise_query_key(tool_name: str, tool_input: str) -> str:
    """Stable key for duplicate-query detection."""
    compact = re.sub(r"\s+", " ", (tool_input or "").strip().lower())
    compact = re.sub(r"[^a-z0-9:/._?#&=% -]", " ", compact)
    compact = re.sub(r"\s+", " ", compact).strip()
    tokens = []
    for token in compact.split(" "):
        if len(token) > 4 and token.endswith(("ing", "ed", "es", "s")):
            token = re.sub(r"(ing|ed|es|s)$", "", token)
        tokens.append(token)
    return f"{tool_name}::{' '.join(tokens)}"


def _build_working_memory_note(
    *,
    evidence_snippets: list[str],
    blocked_query_keys: set[str],
    blocked_hosts: dict[str, int],
    candidate_answer: str | None,
) -> str:
    evidence = evidence_snippets[-3:]
    blocked_q = sorted(blocked_query_keys)[-4:]
    blocked_h = [f"{h}({n})" for h, n in blocked_hosts.items() if n >= 2]
    lines = ["[WORKING MEMORY]"]
    if evidence:
        lines.append("Recent evidence:")
        for item in evidence:
            lines.append(f"- {item}")
    else:
        lines.append("Recent evidence: none yet")
    if blocked_q:
        lines.append("Do-not-repeat queries:")
        for key in blocked_q:
            lines.append(f"- {key}")
    if blocked_h:
        lines.append("Blocked hosts (avoid web_fetch_url): " + ", ".join(sorted(blocked_h)))
    if candidate_answer:
        lines.append(f"Current candidate answer: {candidate_answer}")
    lines.append("Before action=final, verify exact requested output format.")
    return "\n".join(lines)


def _infer_answer_requirements(question: str) -> str:
    q = (question or "").lower()
    rules: list[str] = []
    if "comma-separated" in q or "comma separated" in q:
        rules.append("output comma-separated list with a single space after each comma")
    if "in thousand" in q:
        rules.append("convert to thousands if needed")
    if "round" in q or "rounded" in q:
        rules.append("apply the requested rounding")
    if "name" in q and "full name" not in q:
        rules.append("output name only; no honorifics")
    if "as it appears" in q or "exact phrase" in q:
        rules.append("output exact phrase only; remove scene directives")
    if "number" in q or "how many" in q:
        rules.append("prefer digits-only answer unless units explicitly requested")
    return "; ".join(rules) if rules else "follow literal wording in question"


def _estimate_gaia_confidence(
    *,
    done: bool,
    hedged_final: bool,
    parse_error: bool,
    tool_error: bool,
    duplicate_query: bool,
    tool_error_streak: int,
    dead_end_count: int,
    evidence_count: int,
) -> float:
    """Heuristic confidence for GAIA self-eval routing.

    GAIA runs do not use the text executor's `FINAL ANSWER` loop, so we need
    a native confidence signal in the tool-agent path. This estimate is tuned
    for escalation behavior (routing), not for standalone calibration quality.
    """
    if parse_error:
        return 0.05
    if done:
        if hedged_final:
            return 0.12
        # A concrete final answer with some evidence should be high confidence.
        if evidence_count >= 2 and dead_end_count == 0:
            return 0.86
        if evidence_count >= 1:
            return 0.7
        return 0.45

    score = 0.62
    if tool_error:
        score -= 0.2
    if duplicate_query:
        score -= 0.12
    score -= min(0.3, 0.1 * tool_error_streak)
    score -= min(0.2, 0.03 * dead_end_count)
    if evidence_count == 0:
        score -= 0.08
    return max(0.0, min(1.0, score))


def _reformat_final_answer(
    client: OpenAI,
    model: str,
    question: str,
    candidate: str,
    temperature: float,
    seed: int,
) -> tuple[str, int, int, float]:
    """Run a single cheap LLM pass to coerce `candidate` into GAIA's literal
    format. Returns (reformatted_answer, prompt_tokens, completion_tokens, latency).

    The reformatter MAY only: strip filler/units/articles, re-order lists,
    fix comma spacing, apply rounding/unit conversions explicitly requested
    by the question. It must NOT invent new facts.
    """
    system = (
        "You are a strict answer formatter. Given a question and a candidate "
        "answer, return ONLY the final literal answer as the question asks. "
        "Rules:\n"
        "- Do not add explanations, prefixes, quotes, or trailing punctuation.\n"
        "- For numbers: output digits only (and decimal point). Apply any "
        "rounding or unit conversion the question requests (e.g. 'in thousand "
        "hours' means divide by 1000 and round appropriately). Drop units "
        "unless the question asks for them.\n"
        "- For comma-separated lists: output 'a, b, c' with a space after "
        "each comma; preserve the requested ordering.\n"
        "- For names: output only the name, no honorific, no title.\n"
        "- For an exact phrase asked 'as it appears': strip scene directives "
        "like 'INT.', 'EXT.', leading/trailing '- DAY', '- NIGHT'.\n"
        "- If the candidate already matches the requested format exactly, "
        "return it unchanged.\n"
        "- Never invent information not present in the candidate."
    )
    user = (
        f"Question:\n{question}\n\n"
        f"Candidate answer: {candidate}\n\n"
        "Output only the final literal answer."
    )
    t0 = time.perf_counter()
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user},
            ],
            temperature=temperature,
            seed=seed,
            max_completion_tokens=64,
        )
    except Exception:  # noqa: BLE001
        # On any API issue, keep the candidate as-is.
        return candidate, 0, 0, time.perf_counter() - t0
    latency = time.perf_counter() - t0
    text = (resp.choices[0].message.content or "").strip()
    # Strip any accidental wrapping quotes.
    text = text.strip().strip('"').strip("'").strip()
    if not text:
        text = candidate
    usage = resp.usage
    return (
        text,
        usage.prompt_tokens if usage else 0,
        usage.completion_tokens if usage else 0,
        latency,
    )


def _safe_eval_math(expr: str) -> str:
    """Evaluate arithmetic safely via AST-limited expressions."""
    allowed_nodes = {
        ast.Expression,
        ast.BinOp,
        ast.UnaryOp,
        ast.Add,
        ast.Sub,
        ast.Mult,
        ast.Div,
        ast.FloorDiv,
        ast.Mod,
        ast.Pow,
        ast.USub,
        ast.UAdd,
        ast.Constant,
        ast.Load,
    }

    node = ast.parse(expr, mode="eval")
    for n in ast.walk(node):
        if type(n) not in allowed_nodes:
            raise ValueError("Unsupported expression")
        if isinstance(n, ast.Constant) and not isinstance(n.value, (int, float)):
            raise ValueError("Only numeric constants are allowed")
    result = eval(compile(node, "<calculator>", "eval"), {"__builtins__": {}}, {})
    if isinstance(result, float) and (math.isnan(result) or math.isinf(result)):
        raise ValueError("Invalid numeric result")
    return str(result)


_DEFAULT_HEADERS = {
    "User-Agent": (
        "advisor-eval/1.0 (research benchmark runner; "
        "contact: advisor-eval@example.com)"
    ),
    "Accept": "*/*",
}

_FALLBACK_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

_RETRY_STATUSES = {429, 500, 502, 503, 504}


def _http_request_raw(url: str, timeout: int) -> tuple[bytes, str]:
    """GET a URL with retries, UA rotation, and arxiv mirror fallback.

    Retries transient 5xx/429 once with backoff; on 403 retries once with a
    browser-style UA. For export.arxiv.org / arxiv.org the sibling host is
    tried as a last-ditch fallback.
    """
    attempts: list[tuple[str, dict[str, str]]] = [
        (url, dict(_DEFAULT_HEADERS)),
    ]
    # UA fallback for 403 / JS-gated sites.
    browser_headers = dict(_DEFAULT_HEADERS)
    browser_headers["User-Agent"] = _FALLBACK_UA
    browser_headers["Accept"] = (
        "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    )
    attempts.append((url, browser_headers))
    # arxiv mirror fallback.
    if "export.arxiv.org" in url:
        attempts.append((url.replace("export.arxiv.org", "arxiv.org"), dict(_DEFAULT_HEADERS)))
    elif "://arxiv.org" in url:
        attempts.append((url.replace("://arxiv.org", "://export.arxiv.org"), dict(_DEFAULT_HEADERS)))

    last_exc: Exception | None = None
    for attempt_idx, (attempt_url, headers) in enumerate(attempts):
        for retry_idx in range(2):
            try:
                req = urllib.request.Request(attempt_url, headers=headers, method="GET")
                with urllib.request.urlopen(req, timeout=timeout) as resp:
                    ctype = resp.headers.get("Content-Type", "") or ""
                    return resp.read(), ctype.lower()
            except urllib.error.HTTPError as exc:  # noqa: PERF203
                last_exc = exc
                if exc.code in _RETRY_STATUSES and retry_idx == 0:
                    time.sleep(1.0 + 2.0 * retry_idx)
                    continue
                break  # move to next attempt variant
            except urllib.error.URLError as exc:
                last_exc = exc
                if retry_idx == 0:
                    time.sleep(1.0)
                    continue
                break
            except Exception as exc:  # noqa: BLE001
                last_exc = exc
                break
    assert last_exc is not None
    raise last_exc


def _http_get(url: str) -> str:
    data, _ = _http_request_raw(url, timeout=20)
    return data.decode("utf-8", errors="replace")


def _http_get_bytes(url: str) -> tuple[bytes, str]:
    """Fetch raw bytes plus a best-effort content-type string."""
    return _http_request_raw(url, timeout=30)


def _strip_html(content: str) -> str:
    content = re.sub(r"(?is)<script[^>]*>.*?</script>", " ", content)
    content = re.sub(r"(?is)<style[^>]*>.*?</style>", " ", content)
    content = re.sub(r"(?is)<[^>]+>", " ", content)
    content = html.unescape(content)
    content = re.sub(r"\s+", " ", content).strip()
    return content


def _parse_tool_input(tool_input: str) -> Any:
    payload = tool_input.strip()
    if not payload:
        return ""
    if payload.startswith("{") or payload.startswith("["):
        try:
            return json.loads(payload)
        except json.JSONDecodeError:
            return payload
    return payload


def _as_text(parsed: Any, key: str | None = None) -> str:
    """Extract a plain text query/URL from a parsed tool input (dict or str)."""
    if isinstance(parsed, dict):
        if key and key in parsed:
            return str(parsed.get(key, "")).strip()
        for candidate in ("query", "q", "url", "title", "topic", "text", "input"):
            if candidate in parsed:
                return str(parsed.get(candidate, "")).strip()
        return json.dumps(parsed, ensure_ascii=False)
    return str(parsed).strip()


def _detect_file_kind(name: str, ctype: str = "") -> str:
    name = (name or "").lower()
    ctype = (ctype or "").lower()
    if name.endswith(".pdf") or "pdf" in ctype:
        return "pdf"
    if name.endswith(".xlsx") or name.endswith(".xlsm") or "spreadsheet" in ctype:
        return "xlsx"
    if name.endswith(".xls"):
        return "xls"
    if name.endswith(".docx") or "wordprocessingml" in ctype:
        return "docx"
    if name.endswith(".csv") or "text/csv" in ctype:
        return "csv"
    if (
        name.endswith(".json")
        or name.endswith(".jsonld")
        or "application/json" in ctype
        or "application/ld+json" in ctype
    ):
        return "json"
    if name.endswith(".zip") or "application/zip" in ctype:
        return "zip"
    if name.endswith((".txt", ".md", ".py", ".ipynb", ".tsv", ".log", ".html", ".htm", ".xml")):
        return "text"
    if ctype.startswith("text/"):
        return "text"
    return "binary"


def _extract_pdf_text(data: bytes, max_chars: int = 4000) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except ImportError:
        return "PDF extraction unavailable: pypdf is not installed."
    try:
        reader = PdfReader(io.BytesIO(data))
    except Exception as exc:  # noqa: BLE001
        return f"Failed to read PDF: {exc}"
    out: list[str] = []
    total = 0
    for i, page in enumerate(reader.pages):
        try:
            text = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            text = f"[page {i + 1} extract error: {exc}]"
        if not text:
            continue
        out.append(f"[page {i + 1}]\n{text.strip()}")
        total += len(text)
        if total > max_chars * 2:
            break
    if not out:
        return "PDF contains no extractable text (may be scanned images)."
    joined = "\n\n".join(out)
    return joined[:max_chars]


def _extract_xlsx_text(
    data: bytes,
    sheet: str | None = None,
    max_rows: int = 200,
    max_cols: int = 30,
) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return "XLSX extraction unavailable: openpyxl is not installed."
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return f"Failed to open XLSX: {exc}"
    sheets = wb.sheetnames
    target_sheets: list[str]
    if sheet and sheet in sheets:
        target_sheets = [sheet]
    elif sheet:
        return f"Sheet '{sheet}' not found. Available sheets: {sheets}"
    else:
        target_sheets = sheets[:3]
    chunks: list[str] = [f"sheets={sheets}"]
    for name in target_sheets:
        ws = wb[name]
        chunks.append(f"\n== sheet: {name} ==")
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx >= max_rows:
                chunks.append(f"... ({row_idx} rows shown, more available)")
                break
            trimmed = [
                "" if v is None else str(v)
                for v in row[:max_cols]
            ]
            chunks.append("\t".join(trimmed))
    return "\n".join(chunks)[:6000]


def _extract_docx_text(data: bytes, max_chars: int = 4000) -> str:
    try:
        from docx import Document  # type: ignore
    except ImportError:
        return "DOCX extraction unavailable: python-docx is not installed."
    try:
        doc = Document(io.BytesIO(data))
    except Exception as exc:  # noqa: BLE001
        return f"Failed to open DOCX: {exc}"
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    joined = "\n".join(paragraphs)
    return joined[:max_chars] if joined else "DOCX contains no extractable text."


def _extract_csv_text(data: bytes, max_rows: int = 200) -> str:
    try:
        text = data.decode("utf-8", errors="replace")
    except Exception as exc:  # noqa: BLE001
        return f"Failed to decode CSV: {exc}"
    reader = csv.reader(io.StringIO(text))
    lines: list[str] = []
    for idx, row in enumerate(reader):
        if idx >= max_rows:
            lines.append(f"... ({idx} rows shown, more available)")
            break
        lines.append("\t".join(row))
    return "\n".join(lines)[:6000]


def _extract_zip_text(
    data: bytes,
    member: str | None = None,
    max_chars: int = 4000,
) -> str:
    try:
        zf = zipfile.ZipFile(io.BytesIO(data))
    except Exception as exc:  # noqa: BLE001
        return f"Failed to open ZIP: {exc}"
    members = zf.namelist()
    if not member:
        return f"ZIP contents: {members[:40]}\nCall read_attachment with member='<name>' to read a specific file."
    if member not in members:
        return f"Member '{member}' not found. Contents: {members[:40]}"
    with zf.open(member) as fh:
        raw = fh.read()
    kind = _detect_file_kind(member)
    if kind == "pdf":
        return _extract_pdf_text(raw, max_chars=max_chars)
    if kind == "xlsx":
        return _extract_xlsx_text(raw, max_rows=200)
    if kind == "docx":
        return _extract_docx_text(raw, max_chars=max_chars)
    if kind == "csv":
        return _extract_csv_text(raw)
    try:
        return raw.decode("utf-8", errors="replace")[:max_chars]
    except Exception as exc:  # noqa: BLE001
        return f"Failed to decode member '{member}': {exc}"


def _extract_by_kind(
    data: bytes,
    kind: str,
    options: dict[str, Any] | None = None,
) -> str:
    opts = options or {}
    if kind == "pdf":
        return _extract_pdf_text(data, max_chars=int(opts.get("max_chars", 4000)))
    if kind == "xlsx":
        return _extract_xlsx_text(
            data,
            sheet=opts.get("sheet"),
            max_rows=int(opts.get("max_rows", 200)),
            max_cols=int(opts.get("max_cols", 30)),
        )
    if kind == "docx":
        return _extract_docx_text(data, max_chars=int(opts.get("max_chars", 4000)))
    if kind == "csv":
        return _extract_csv_text(data, max_rows=int(opts.get("max_rows", 200)))
    if kind == "zip":
        return _extract_zip_text(
            data,
            member=opts.get("member"),
            max_chars=int(opts.get("max_chars", 4000)),
        )
    if kind == "json":
        try:
            return data.decode("utf-8", errors="replace")[:4000]
        except Exception as exc:  # noqa: BLE001
            return f"Failed to decode JSON: {exc}"
    if kind == "text":
        try:
            return data.decode("utf-8", errors="replace")[:4000]
        except Exception as exc:  # noqa: BLE001
            return f"Failed to decode text: {exc}"
    # Fallback: try utf-8 then report binary size.
    try:
        decoded = data.decode("utf-8")
        return decoded[:4000]
    except Exception:  # noqa: BLE001
        return (
            f"Binary content ({len(data)} bytes) that could not be decoded. "
            "If this is an attachment, consider requesting a different format."
        )


def _wiki_search(parsed: Any) -> str:
    query = _as_text(parsed, key="query")
    if not query:
        raise ValueError("wiki_search requires a non-empty query string.")
    params = urllib.parse.urlencode({
        "action": "query",
        "list": "search",
        "srsearch": query,
        "format": "json",
        "srlimit": 3,
    })
    url = f"https://en.wikipedia.org/w/api.php?{params}"
    payload = json.loads(_http_get(url))
    items = payload.get("query", {}).get("search", [])
    if not items:
        return "No results."
    rows = []
    for idx, item in enumerate(items, 1):
        title = item.get("title", "")
        snippet = item.get("snippet", "").replace("<span class=\"searchmatch\">", "").replace("</span>", "")
        rows.append(f"{idx}. {title}: {snippet}")
    return "\n".join(rows)


def _wiki_lookup(parsed: Any) -> str:
    topic = _as_text(parsed, key="title") or _as_text(parsed, key="topic")
    if not topic:
        raise ValueError("wiki_lookup requires a page title string.")
    params = urllib.parse.urlencode({
        "action": "query",
        "prop": "extracts",
        "explaintext": 1,
        "exintro": 1,
        "titles": topic,
        "format": "json",
    })
    url = f"https://en.wikipedia.org/w/api.php?{params}"
    payload = json.loads(_http_get(url))
    pages = payload.get("query", {}).get("pages", {})
    if not pages:
        return "No page found."
    page = next(iter(pages.values()))
    extract = page.get("extract", "")
    if not extract:
        return "No extract available."
    return extract[:2000]


def _ddg_fetch_matches(query: str) -> list[tuple[str, str]]:
    q = urllib.parse.quote_plus(query)
    endpoints = [
        f"https://duckduckgo.com/html/?q={q}",
        f"https://html.duckduckgo.com/html/?q={q}",
    ]
    for url in endpoints:
        try:
            page = _http_get(url)
        except Exception:  # noqa: BLE001
            continue
        matches = re.findall(
            r'(?is)<a[^>]*class="[^"]*result__a[^"]*"[^>]*href="([^"]+)"[^>]*>(.*?)</a>',
            page,
        )
        if matches:
            return matches
    return []


def _web_search(parsed: Any) -> str:
    query = _as_text(parsed, key="query")
    if not query:
        raise ValueError("web_search requires a non-empty query string.")
    matches = _ddg_fetch_matches(query)
    # If a site: query returns nothing, retry without the site: prefix.
    if not matches and re.search(r"\bsite:\S+", query):
        relaxed = re.sub(r"\bsite:\S+\s*", "", query).strip()
        if relaxed and relaxed != query:
            matches = _ddg_fetch_matches(relaxed)
    if not matches:
        return "No results."
    rows = []
    for idx, (href, title_html) in enumerate(matches[:5], 1):
        title = _strip_html(title_html)
        # DuckDuckGo wraps hrefs; try to unwrap uddg=...
        if "uddg=" in href:
            try:
                qs = urllib.parse.urlparse(href).query
                params = urllib.parse.parse_qs(qs)
                if "uddg" in params:
                    href = urllib.parse.unquote(params["uddg"][0])
            except Exception:  # noqa: BLE001
                pass
        # DDG lite prefixes hrefs with // - normalise.
        if href.startswith("//"):
            href = "https:" + href
        rows.append(f"{idx}. {title} - {href}")
    return "\n".join(rows)


def _web_fetch_url(parsed: Any) -> str:
    options: dict[str, Any] = {}
    if isinstance(parsed, dict):
        url = str(parsed.get("url", "")).strip()
        options = {k: v for k, v in parsed.items() if k != "url"}
    else:
        url = str(parsed).strip()
    if not url:
        raise ValueError("web_fetch_url requires a URL.")
    parsed_url = urllib.parse.urlparse(url)
    if parsed_url.scheme not in {"http", "https"}:
        raise ValueError("URL must start with http:// or https://")
    data, ctype = _http_get_bytes(url)
    guessed_name = Path(parsed_url.path).name
    if not mimetypes.guess_type(guessed_name)[0] and not ctype:
        ctype = "text/html"
    kind = _detect_file_kind(guessed_name, ctype)
    if kind in {"pdf", "xlsx", "docx", "csv", "zip", "json"}:
        return _extract_by_kind(data, kind, options)
    # Treat text/html (and unknown) as HTML to strip.
    try:
        content = data.decode("utf-8", errors="replace")
    except Exception as exc:  # noqa: BLE001
        return f"Failed to decode URL content: {exc}"
    if "<html" in content.lower() or "<!doctype" in content.lower() or "<body" in content.lower():
        text = _strip_html(content)
        return text[:3500] if text else "No readable text found."
    return content[:3500]


def _arxiv_search(parsed: Any) -> str:
    query = _as_text(parsed, key="query").strip()
    if not query:
        raise ValueError("arxiv_search requires a query string.")
    # Accept both plain-text and fielded arXiv queries.
    # Example fielded query:
    #   cat:physics.soc-ph AND submittedDate:[201608110000 TO 201608120000]
    fielded_tokens = ("cat:", "ti:", "abs:", "au:", "submittedDate:", "id:")
    if any(tok in query for tok in fielded_tokens):
        search_query = query
    else:
        search_query = f"all:{query}"

    # Convenience: convert submittedDate:YYYY-MM-DD into a day range.
    date_match = re.search(r"submittedDate:(\d{4})-(\d{2})-(\d{2})", search_query)
    if date_match:
        y, m, d = date_match.groups()
        dt = datetime(int(y), int(m), int(d))
        end_dt = dt + timedelta(days=1)
        start = dt.strftime("%Y%m%d0000")
        end = end_dt.strftime("%Y%m%d0000")
        search_query = re.sub(
            r"submittedDate:\d{4}-\d{2}-\d{2}",
            f"submittedDate:[{start} TO {end}]",
            search_query,
        )

    encoded = urllib.parse.quote_plus(search_query)
    url = (
        "https://export.arxiv.org/api/query"
        f"?search_query={encoded}&start=0&max_results=5"
    )
    xml_text = _http_get(url)
    ns = {"atom": "http://www.w3.org/2005/Atom"}
    root = ET.fromstring(xml_text)
    entries = root.findall("atom:entry", ns)
    if not entries:
        return "No arXiv results."
    rows = []
    for idx, entry in enumerate(entries, 1):
        title = (entry.findtext("atom:title", default="", namespaces=ns) or "").strip()
        arxiv_id = (entry.findtext("atom:id", default="", namespaces=ns) or "").strip()
        published = (entry.findtext("atom:published", default="", namespaces=ns) or "").strip()
        summary = (entry.findtext("atom:summary", default="", namespaces=ns) or "").strip()
        rows.append(
            f"{idx}. {title}\n"
            f"   id: {arxiv_id}\n"
            f"   published: {published}\n"
            f"   summary: {summary[:220]}"
        )
    return "\n".join(rows)


def _arxiv_fetch(parsed: Any) -> str:
    identifier = _as_text(parsed, key="id") or _as_text(parsed, key="arxiv_id")
    identifier = identifier.strip()
    if not identifier:
        raise ValueError("arxiv_fetch requires an arXiv id.")
    match = re.search(r"(\d{4}\.\d{4,5}(v\d+)?)", identifier)
    if match:
        arxiv_id = match.group(1)
    else:
        arxiv_id = identifier
    url = f"https://export.arxiv.org/api/query?id_list={urllib.parse.quote_plus(arxiv_id)}"
    xml_text = _http_get(url)
    ns = {"atom": "http://www.w3.org/2005/Atom"}
    root = ET.fromstring(xml_text)
    entry = root.find("atom:entry", ns)
    if entry is None:
        return "No arXiv entry found."
    title = (entry.findtext("atom:title", default="", namespaces=ns) or "").strip()
    published = (entry.findtext("atom:published", default="", namespaces=ns) or "").strip()
    authors = [
        (a.findtext("atom:name", default="", namespaces=ns) or "").strip()
        for a in entry.findall("atom:author", ns)
    ]
    summary = (entry.findtext("atom:summary", default="", namespaces=ns) or "").strip()
    return (
        f"title: {title}\n"
        f"published: {published}\n"
        f"authors: {', '.join([a for a in authors if a])}\n"
        f"summary: {summary[:2400]}"
    )


def _github_issue_fetch(parsed: Any) -> str:
    if isinstance(parsed, dict):
        repo = str(parsed.get("repo", "")).strip()
        issue_number = str(parsed.get("issue_number", "")).strip()
    else:
        raw = str(parsed).strip()
        # Accept "owner/repo#123" or a full github URL.
        url_match = re.match(
            r"https?://github\.com/([^/]+/[^/]+)/(?:issues|pull)/(\d+)", raw
        )
        if url_match:
            repo, issue_number = url_match.group(1), url_match.group(2)
        else:
            parts = raw.split("#")
            if len(parts) != 2:
                raise ValueError(
                    "Use JSON {'repo':'owner/repo','issue_number':123} or 'owner/repo#123'."
                )
            repo, issue_number = parts[0].strip(), parts[1].strip()
    if not repo or not issue_number:
        raise ValueError("Missing repo or issue_number")
    base = f"https://api.github.com/repos/{repo}/issues/{issue_number}"
    issue = json.loads(_http_get(base))
    labels = [lbl.get("name", "") for lbl in issue.get("labels", []) if isinstance(lbl, dict)]
    timeline_url = f"{base}/timeline"
    try:
        req = urllib.request.Request(
            timeline_url,
            headers={**_DEFAULT_HEADERS, "Accept": "application/vnd.github+json"},
            method="GET",
        )
        with urllib.request.urlopen(req, timeout=20) as resp:
            timeline = json.loads(resp.read().decode("utf-8", errors="replace"))
    except Exception:  # noqa: BLE001
        timeline = []
    label_events: list[str] = []
    for event in timeline:
        event_type = event.get("event", "")
        if event_type in {"labeled", "unlabeled", "closed", "reopened"}:
            label = event.get("label", {}).get("name", "") if isinstance(event.get("label"), dict) else ""
            created_at = event.get("created_at", "")
            label_events.append(f"{event_type}:{label}@{created_at}")
    body = (issue.get("body") or "").strip().replace("\r\n", "\n")
    if len(body) > 1200:
        body = body[:1200] + "..."
    return (
        f"title: {issue.get('title','')}\n"
        f"state: {issue.get('state','')}\n"
        f"created_at: {issue.get('created_at','')}\n"
        f"closed_at: {issue.get('closed_at','')}\n"
        f"user: {issue.get('user', {}).get('login','') if isinstance(issue.get('user'), dict) else ''}\n"
        f"labels: {', '.join(labels)}\n"
        f"timeline_events: {', '.join(label_events[:20])}\n"
        f"body: {body}"
    )


def _read_attachment(parsed: Any, task_metadata: dict[str, Any]) -> str:
    file_name = task_metadata.get("file_name", "") or ""
    file_path = task_metadata.get("file_path", "") or ""
    dataset_id = task_metadata.get("dataset_id", "gaia-benchmark/GAIA")
    dataset_config = task_metadata.get("dataset_config_name") or task_metadata.get("dataset_config")

    options: dict[str, Any] = {}
    if isinstance(parsed, dict):
        if parsed.get("file_path"):
            file_path = str(parsed["file_path"])
        if parsed.get("file_name"):
            file_name = str(parsed["file_name"])
        options = {k: v for k, v in parsed.items() if k not in {"file_path", "file_name"}}

    if not file_path and not file_name:
        return "No attachment metadata available for this task."

    candidate = file_path or file_name
    kind = _detect_file_kind(file_name or candidate)

    local_path: Path | None = None
    if Path(candidate).exists():
        local_path = Path(candidate)

    if local_path is None:
        # Try resolving artifact from HF dataset repository.
        tried_paths: list[str] = []
        for filename_candidate in (
            candidate,
            file_name,
            f"{dataset_config}/{file_name}" if dataset_config and file_name else None,
            f"2023/validation/{file_name}" if file_name else None,
            f"2023/test/{file_name}" if file_name else None,
        ):
            if not filename_candidate or filename_candidate in tried_paths:
                continue
            tried_paths.append(filename_candidate)
            try:
                local_fp = hf_hub_download(
                    repo_id=dataset_id,
                    repo_type="dataset",
                    filename=filename_candidate,
                )
                local_path = Path(local_fp)
                break
            except Exception:  # noqa: BLE001
                continue

    if local_path is None:
        return (
            "Attachment could not be loaded automatically.\n"
            f"file_name={file_name}\nfile_path={file_path}\n"
            "Hint: pass {'file_name':'...'} as tool_input if metadata is missing."
        )

    data = local_path.read_bytes()
    effective_kind = kind or _detect_file_kind(local_path.name)
    if effective_kind == "binary":
        effective_kind = _detect_file_kind(local_path.name)
    extracted = _extract_by_kind(data, effective_kind, options)
    header = (
        f"[attachment: {local_path.name}, kind={effective_kind}, "
        f"size={len(data)} bytes]\n"
    )
    return header + extracted


def _truncate_messages_for_advisor(
    messages: list[dict[str, str]],
    per_msg_cap: int = 1200,
) -> list[dict[str, str]]:
    """Cap very long tool-result messages to keep the advisor focused."""
    out: list[dict[str, str]] = []
    for m in messages:
        content = m.get("content", "")
        if isinstance(content, str) and len(content) > per_msg_cap and (
            content.startswith("[TOOL RESULT]") or content.startswith("[TOOL ERROR]")
        ):
            content = content[: per_msg_cap - 40] + "\n…[truncated for advisor]"
        out.append({"role": m.get("role", "user"), "content": content})
    return out


def _build_advisor_context_pack(
    *,
    trigger: str,
    question: str,
    attachment_hint: str,
    tool_trace: list[dict[str, Any]],
    last_raw: str,
    last_action: dict[str, Any] | None,
    budget_used: int,
    budget_max: int,
    blocked_hosts: dict[str, int],
    blocked_query_keys: set[str],
    evidence_snippets: list[str],
    candidate_answer: str | None,
    tool_error_streak: int,
    dead_end_count: int,
) -> str:
    lines: list[str] = [f"[ADVISOR CONSULT: trigger={trigger}]"]
    lines.append(f"Question: {question.strip()}")
    lines.append(f"Answer requirements: {_infer_answer_requirements(question)}")
    lines.append(
        "Available tools: wiki_search, wiki_lookup, web_search, web_fetch_url, "
        "arxiv_search, arxiv_fetch, github_issue_fetch, read_attachment, calculator"
    )
    if attachment_hint:
        lines.append(f"Attachment: {attachment_hint}")
    else:
        lines.append("Attachment: none")
    lines.append(f"Tool budget: {budget_used}/{budget_max}")
    lines.append(
        f"Health signals: tool_error_streak={tool_error_streak}, dead_end_count={dead_end_count}"
    )
    if candidate_answer:
        lines.append(f"Current candidate answer: {_short(candidate_answer, 160)}")
    if evidence_snippets:
        lines.append("Recent evidence snippets:")
        for snippet in evidence_snippets[-3:]:
            lines.append(f"  - {snippet}")
    blocked_hosts_list = [f"{h}({n})" for h, n in blocked_hosts.items() if n >= 2]
    if blocked_hosts_list:
        lines.append("Blocked hosts (avoid refetch): " + ", ".join(sorted(blocked_hosts_list)))
    if blocked_query_keys:
        lines.append("Blocked duplicate queries:")
        for item in sorted(blocked_query_keys)[-5:]:
            lines.append(f"  - {item}")
    # Compact numbered tool history.
    if tool_trace:
        lines.append(f"Tools used so far (n={len(tool_trace)}):")
        for idx, entry in enumerate(tool_trace[-10:], 1):
            tool = entry.get("tool", "?")
            inp = _short(str(entry.get("input", "")), 80)
            if entry.get("success"):
                preview = _short(str(entry.get("output_preview", "")), 100)
                lines.append(f"  {idx}. {tool}(\"{inp}\") -> ok: {preview}")
            else:
                err = _short(str(entry.get("error", "")), 100)
                lines.append(f"  {idx}. {tool}(\"{inp}\") -> error: {err}")
    else:
        lines.append("Tools used so far: none")
    if last_action is not None:
        act = str(last_action.get("action", "?"))
        if act == "final":
            lines.append(
                f"Executor's last step: action=final, final_answer="
                f"\"{_short(str(last_action.get('final_answer','')), 120)}\""
            )
        else:
            lines.append(
                f"Executor's last step: action=tool, "
                f"tool_name={last_action.get('tool_name','?')}, "
                f"tool_input=\"{_short(str(last_action.get('tool_input','')), 120)}\""
            )
    elif last_raw:
        lines.append(f"Executor's last raw output: {_short(last_raw, 200)}")
    lines.append("Give DIAGNOSIS + NEXT + AVOID now.")
    return "\n".join(lines)


def _parse_next_action_from_guidance(guidance: str) -> tuple[str | None, str | None]:
    """Extract first tool and a best-effort input hint from advisor NEXT."""
    if not guidance:
        return None, None
    # Find the NEXT: section.
    m = re.search(r"(?is)\bNEXT\s*:\s*(.*?)(?:\bAVOID\s*:|$)", guidance)
    block = m.group(1) if m else guidance
    tools = (
        "wiki_search",
        "wiki_lookup",
        "web_search",
        "web_fetch_url",
        "arxiv_search",
        "arxiv_fetch",
        "github_issue_fetch",
        "read_attachment",
        "calculator",
    )
    for line in block.splitlines():
        line = line.strip()
        # Skip empty / heading lines.
        if not line:
            continue
        for tool in tools:
            if re.search(rf"(?i)`?{tool}`?\b", line):
                input_match = re.search(
                    r'(?i)(?:input|query|url)\s*[:=]\s*["“]?([^"\n”]+)',
                    line,
                )
                return tool, (input_match.group(1).strip() if input_match else None)
    # Fall back to any tool mentioned anywhere in guidance.
    for tool in tools:
        if re.search(rf"(?i)`?{tool}`?\b", guidance):
                return tool, None
    return None, None


def _guidance_is_actionable(
    guidance: str,
    tool_trace: list[dict[str, Any]],
    blocked_hosts: dict[str, int],
) -> bool:
    """Reject weak advisor outputs that repeat known-bad paths."""
    if not guidance or "NEXT:" not in guidance:
        return False
    lower = guidance.lower()
    # Reject if guidance asks to revisit hosts already blocked 2+ times.
    for host, count in blocked_hosts.items():
        if count >= 2 and host and host in lower:
            return False
    # Reject exact repeated failing (tool,input) hints already seen.
    failed_pairs = {
        (
            str(entry.get("tool", "")).strip().lower(),
            str(entry.get("input", "")).strip().lower(),
        )
        for entry in tool_trace
        if not entry.get("success", True)
    }
    for tool, input_text in failed_pairs:
        if tool and input_text and tool in lower and input_text[:80] in lower:
            return False
    return True


def _extract_host(tool_name: str, tool_input: str) -> str | None:
    """Return a bare hostname for web_fetch_url calls (for blocked-host tracking)."""
    if tool_name != "web_fetch_url":
        return None
    try:
        from urllib.parse import urlparse
        parsed = urlparse(tool_input.strip().strip('"\''))
        host = (parsed.netloc or "").lower()
        # Strip user-info, port, leading "www.".
        if "@" in host:
            host = host.split("@", 1)[1]
        if ":" in host:
            host = host.split(":", 1)[0]
        if host.startswith("www."):
            host = host[4:]
        return host or None
    except Exception:  # noqa: BLE001
        return None


def run_gaia_agentic(
    *,
    question: str,
    executor_model: str,
    advisor: AdvisorAgent | None,
    policy: EscalationPolicy,
    policy_name: str,
    temperature: float,
    seed: int,
    max_steps: int,
    max_tool_calls: int,
    max_advisor_calls: int = 2,
    task_metadata: dict[str, Any] | None = None,
) -> GaiaRunResult:
    """Run an agentic loop with tool usage and optional advisor support."""
    client = OpenAI()
    result = GaiaRunResult()
    had_error = False
    post_error_recovered = False

    task_metadata = task_metadata or {}
    attachment_info = []
    file_name = task_metadata.get("file_name") or ""
    attachment_present = bool(file_name or task_metadata.get("file_path"))
    if file_name:
        attachment_info.append(f"attachment_file_name={file_name}")
        attachment_info.append(
            f"attachment_kind={_detect_file_kind(file_name)}"
        )
    if task_metadata.get("file_path"):
        attachment_info.append(f"attachment_file_path={task_metadata.get('file_path')}")
    if task_metadata.get("level"):
        attachment_info.append(f"gaia_level={task_metadata.get('level')}")
    if attachment_present:
        attachment_hint = (
            "\n".join(attachment_info)
            + "\nIMPORTANT: call read_attachment FIRST to inspect the attached file "
            "(pass an empty tool_input to read the default file, or JSON with "
            "'sheet'/'max_rows' for XLSX or 'member' for ZIP)."
        )
    else:
        # Explicitly tell the executor not to waste a call on read_attachment.
        attachment_hint = (
            "No attachment for this task. Do NOT call read_attachment; "
            "rely on web/wiki/arxiv tools instead."
        )
    attachment_summary = (
        f"{file_name} ({_detect_file_kind(file_name)})" if file_name else ""
    )

    messages: list[dict[str, str]] = [
        {"role": "system", "content": GAIA_SYSTEM_PROMPT},
        {"role": "user", "content": f"{question}\n\n[Task metadata]\n{attachment_hint}"},
    ]

    # Stuck-detection and advisor cooldown state.
    recent_query_keys: list[str] = []
    blocked_query_keys: set[str] = set()
    evidence_snippets: list[str] = []
    candidate_answer: str | None = None
    tool_error_streak = 0
    cooldown_until = -1
    last_action_dict: dict[str, Any] | None = None
    force_answer_used = False
    late_rescue_used = False
    no_progress_advisor_rounds = 0
    # Hosts we have already seen fail >=2 times. We skip advisor escalation
    # triggered purely by more errors on the same host -- the advisor can't fix
    # Cloudflare/403/anti-bot gates and would just burn the budget.
    blocked_hosts: dict[str, int] = {}
    last_error_host: str | None = None

    for step_idx in range(max_steps):
        step_lat = 0.0
        t0 = time.perf_counter()
        try:
            working_memory_note = _build_working_memory_note(
                evidence_snippets=evidence_snippets,
                blocked_query_keys=blocked_query_keys,
                blocked_hosts=blocked_hosts,
                candidate_answer=candidate_answer,
            )
            response = client.chat.completions.create(
                model=executor_model,
                messages=messages + [{"role": "user", "content": working_memory_note}],
                temperature=temperature,
                seed=seed,
                max_completion_tokens=512,
            )
        except Exception as exc:  # noqa: BLE001
            # OpenAI may reject a prompt (e.g. safety filter) or throttle us.
            # Record the failure and stop the loop gracefully so the matrix
            # continues to the next task instead of crashing.
            latency = time.perf_counter() - t0
            result.total_exec_latency += latency
            result.tool_trace.append({
                "step": step_idx,
                "tool": "executor_api",
                "input": "",
                "success": False,
                "error": f"{type(exc).__name__}: {exc}"[:400],
            })
            result.step_latencies.append(latency)
            result.dead_end_count += 1
            result.recovery_success = False
            return result
        latency = time.perf_counter() - t0
        step_lat += latency

        usage = response.usage
        result.total_exec_latency += latency
        result.total_exec_prompt += usage.prompt_tokens if usage else 0
        result.total_exec_completion += usage.completion_tokens if usage else 0

        raw_text = response.choices[0].message.content or ""
        messages.append({"role": "assistant", "content": raw_text})

        action = _extract_json(raw_text)
        parse_error = action is None
        tool_error = False
        done = False
        wants_advisor = False
        answer: str | None = None
        duplicate_query = False

        if parse_error:
            result.dead_end_count += 1
            tool_error = True
            had_error = True
            tool_error_streak += 1
            feedback = (
                "Your previous response was not valid JSON. "
                "Return only a valid JSON object following the schema."
            )
            messages.append({"role": "user", "content": feedback})
            result.tool_trace.append({
                "step": step_idx,
                "tool": "json_parse",
                "input": "",
                "success": False,
                "error": "invalid_json",
                "output_preview": (raw_text or "")[:400],
            })
        else:
            action_type = str(action.get("action", "")).strip().lower()
            # Some models emit shorthand actions like {"action": "calculator", ...}.
            # Treat those as tool calls to reduce avoidable dead-ends.
            if action_type in {
                "wiki_search",
                "wiki_lookup",
                "web_search",
                "web_fetch_url",
                "arxiv_search",
                "arxiv_fetch",
                "github_issue_fetch",
                "read_attachment",
                "calculator",
            }:
                action["tool_name"] = action_type
                action_type = "tool"
            wants_advisor = bool(action.get("request_advisor", False))
            last_action_dict = {
                "action": action_type,
                "tool_name": action.get("tool_name", ""),
                "tool_input": action.get("tool_input", ""),
                "final_answer": action.get("final_answer", ""),
            }

            if action_type == "final":
                done = True
                answer = str(action.get("final_answer", "")).strip() or None
                candidate_answer = answer
                tool_error_streak = 0
            elif action_type == "tool":
                if result.tool_calls >= max_tool_calls:
                    tool_error = True
                    had_error = True
                    result.dead_end_count += 1
                    messages.append({
                        "role": "user",
                        "content": (
                            f"Tool budget exceeded ({max_tool_calls}). "
                            "Decide with available evidence and return action=final."
                        ),
                    })
                else:
                    tool_name = str(action.get("tool_name", "")).strip()
                    tool_input = str(action.get("tool_input", "")).strip()
                    parsed_tool_input = _parse_tool_input(tool_input)
                    # Duplicate-query detection (signal for advisor trigger).
                    qkey = _normalise_query_key(tool_name, tool_input)
                    if qkey in recent_query_keys:
                        duplicate_query = True
                    if qkey in blocked_query_keys:
                        duplicate_query = True
                        result.repeated_query_violations += 1
                        tool_error = True
                        had_error = True
                        tool_error_streak += 1
                        result.dead_end_count += 1
                        messages.append({
                            "role": "user",
                            "content": (
                                f"[TOOL ERROR] name={tool_name}\ninput={tool_input}\n"
                                "error=duplicate_or_blocked_query\n"
                                "Use a materially different query or switch tools."
                            ),
                        })
                    if qkey not in blocked_query_keys:
                        recent_query_keys.append(qkey)
                        if len(recent_query_keys) > 6:
                            recent_query_keys = recent_query_keys[-6:]
                        result.tool_calls += 1
                        success = True
                        output = ""
                        error = ""
                        try:
                            if tool_name == "wiki_search":
                                output = _wiki_search(parsed_tool_input)
                            elif tool_name == "wiki_lookup":
                                output = _wiki_lookup(parsed_tool_input)
                            elif tool_name == "web_search":
                                output = _web_search(parsed_tool_input)
                            elif tool_name == "web_fetch_url":
                                output = _web_fetch_url(parsed_tool_input)
                            elif tool_name == "arxiv_search":
                                output = _arxiv_search(parsed_tool_input)
                            elif tool_name == "arxiv_fetch":
                                output = _arxiv_fetch(parsed_tool_input)
                            elif tool_name == "github_issue_fetch":
                                output = _github_issue_fetch(parsed_tool_input)
                            elif tool_name == "read_attachment":
                                output = _read_attachment(parsed_tool_input, task_metadata)
                            elif tool_name == "calculator":
                                expr = _as_text(parsed_tool_input, key="expression")
                                output = _safe_eval_math(expr or tool_input)
                            else:
                                raise ValueError(f"Unknown tool '{tool_name}'")
                        except Exception as exc:  # noqa: BLE001
                            success = False
                            error = str(exc)
                            output = ""
                    else:
                        success = False
                        output = ""
                        error = "duplicate_or_blocked_query"

                    if success:
                        if had_error:
                            post_error_recovered = True
                        tool_error_streak = 0
                        messages.append({
                            "role": "user",
                            "content": (
                                f"[TOOL RESULT] name={tool_name}\n"
                                f"input={tool_input}\n"
                                f"output={output}"
                            ),
                        })
                        evidence_snippets.append(f"{tool_name}: {_short(output, 110)}")
                    else:
                        tool_error = True
                        had_error = True
                        tool_error_streak += 1
                        result.tool_errors += 1
                        result.dead_end_count += 1
                        host = _extract_host(tool_name, tool_input)
                        last_error_host = host
                        if host:
                            blocked_hosts[host] = blocked_hosts.get(host, 0) + 1
                            if blocked_hosts[host] >= 2:
                                blocked_query_keys.add(qkey)
                            if blocked_hosts[host] > 2:
                                result.blocked_host_rehits += 1
                        else:
                            blocked_query_keys.add(qkey)
                        extra_hint = ""
                        if host and blocked_hosts.get(host, 0) >= 2:
                            extra_hint = (
                                f" Host {host} has failed {blocked_hosts[host]} "
                                "times -- stop fetching it and use web_search or "
                                "wiki_search instead."
                            )
                        messages.append({
                            "role": "user",
                            "content": (
                                f"[TOOL ERROR] name={tool_name}\n"
                                f"input={tool_input}\n"
                                f"error={error}\n"
                                f"Adjust tool selection or input.{extra_hint}"
                            ),
                        })

                    result.tool_trace.append({
                        "step": step_idx,
                        "tool": tool_name,
                        "input": tool_input,
                        "success": success,
                        "error": error,
                        "output_preview": output[:240],
                    })
            else:
                tool_error = True
                had_error = True
                result.dead_end_count += 1
                messages.append({
                    "role": "user",
                    "content": (
                        "Invalid action type. Use action='tool' or action='final' in valid JSON."
                    ),
                })
                result.tool_trace.append({
                    "step": step_idx,
                    "tool": "action_validation",
                    "input": str(action.get("action", "")),
                    "success": False,
                    "error": "invalid_action",
                })

        hedged_final = done and _is_hedged_answer(answer)
        budget_fraction = (
            result.tool_calls / max_tool_calls if max_tool_calls > 0 else 0.0
        )
        confidence = _estimate_gaia_confidence(
            done=done,
            hedged_final=hedged_final,
            parse_error=parse_error,
            tool_error=tool_error,
            duplicate_query=duplicate_query,
            tool_error_streak=tool_error_streak,
            dead_end_count=result.dead_end_count,
            evidence_count=len(evidence_snippets),
        )
        policy_result: dict[str, Any] = {
            "text": raw_text,
            "answer": answer,
            "wants_advisor": wants_advisor,
            "done": done,
            "tool_error": tool_error,
            "parse_error": parse_error,
            "duplicate_query": duplicate_query,
            "hedged_final": hedged_final,
            "tool_error_streak": tool_error_streak,
            "budget_fraction": budget_fraction,
            "confidence": confidence,
        }
        policy_state: dict[str, Any] = {
            "step": step_idx,
            "messages": messages,
            "dead_end_count": result.dead_end_count,
            "tool_errors": result.tool_errors,
            "tool_error_streak": tool_error_streak,
            "duplicate_query_streak": sum(
                1 for k in recent_query_keys[-2:] if recent_query_keys.count(k) >= 2
            ),
            "budget_fraction": budget_fraction,
            "hedged_final": hedged_final,
        }

        # Determine advisor trigger (StuckPolicy OR configured policy).
        stuck_triggers: list[str] = []
        if tool_error_streak >= 2:
            stuck_triggers.append("two_tool_errors")
        if duplicate_query:
            stuck_triggers.append("duplicate_query")
        if budget_fraction >= 0.75 and not done:
            stuck_triggers.append("budget_near_limit")
        if hedged_final:
            stuck_triggers.append("hedged_final_answer")
        if done and len(evidence_snippets) < 2:
            stuck_triggers.append("low_evidence_final")
        if done and not wants_advisor and (
            hedged_final or len(evidence_snippets) < 2 or result.tool_errors > 0
        ):
            stuck_triggers.append("auto_low_confidence_final")
        if wants_advisor:
            stuck_triggers.append("model_requested")
        if parse_error:
            stuck_triggers.append("parse_error")

        # If the only stuck trigger is "two_tool_errors" on a host we've
        # already marked blocked, suppress it: the advisor can't unblock the
        # site, and an advisor round would just burn budget. Let the regular
        # host-blocked hint above steer the executor to a different tool.
        if (
            stuck_triggers == ["two_tool_errors"]
            and last_error_host is not None
            and blocked_hosts.get(last_error_host, 0) >= 2
        ):
            stuck_triggers = []

        policy_triggered = advisor is not None and policy.should_escalate(
            step_idx, policy_result, policy_state
        )
        stuck_triggered = advisor is not None and bool(stuck_triggers)

        # Hard cap on advisor calls per task.
        advisor_cap_hit = (
            max_advisor_calls is not None
            and result.advisor_calls >= max_advisor_calls
        )
        late_rescue = done and (hedged_final or len(evidence_snippets) < 2)
        # Budget guard: don't escalate when too little tool budget remains
        # to act on advisor guidance (unless we're on a hedged final).
        tool_budget_left = max_tool_calls - result.tool_calls
        budget_guard_block = tool_budget_left < 3 and not hedged_final

        # Random_prob suppression in the first 2 steps (no info yet) and
        # last 3 tool-budget steps (can't act on advice anyway).
        suppress_random = (
            policy_name == "random_prob"
            and (step_idx < 2 or tool_budget_left <= 3)
            and not stuck_triggered
            and not hedged_final
        )
        if suppress_random:
            policy_triggered = False

        # Keep a hard advisor-call cap. Allow one late-rescue bypass only once.
        if late_rescue and late_rescue_used:
            late_rescue = False
        in_cooldown = step_idx <= cooldown_until and not late_rescue
        blocked = advisor_cap_hit or (budget_guard_block and not late_rescue)
        should_escalate = (
            (policy_triggered or stuck_triggered)
            and not in_cooldown
            and not blocked
        )

        if should_escalate and advisor is not None:
            if stuck_triggers:
                trigger_name = stuck_triggers[0]
            else:
                trigger_name = policy_name or "policy"
            context_pack = _build_advisor_context_pack(
                trigger=trigger_name,
                question=question,
                attachment_hint=attachment_summary,
                tool_trace=result.tool_trace,
                last_raw=raw_text,
                last_action=last_action_dict,
                budget_used=result.tool_calls,
                budget_max=max_tool_calls,
                blocked_hosts=blocked_hosts,
                blocked_query_keys=blocked_query_keys,
                evidence_snippets=evidence_snippets,
                candidate_answer=candidate_answer,
                tool_error_streak=tool_error_streak,
                dead_end_count=result.dead_end_count,
            )
            advisor_messages = _truncate_messages_for_advisor(messages) + [
                {"role": "user", "content": context_pack}
            ]
            try:
                guidance, adv_stats = advisor.advise(advisor_messages)
            except Exception as exc:  # noqa: BLE001
                result.tool_trace.append({
                    "step": step_idx,
                    "tool": "advisor_api",
                    "input": trigger_name,
                    "success": False,
                    "error": f"{type(exc).__name__}: {exc}"[:400],
                })
                guidance = None
                adv_stats = None
            if guidance is not None and adv_stats is not None:
                if not _guidance_is_actionable(guidance, result.tool_trace, blocked_hosts):
                    guidance = None
                    adv_stats = None
            if guidance is not None and adv_stats is not None:
                result.total_adv_latency += adv_stats.latency_s
                result.total_adv_prompt += adv_stats.prompt_tokens
                result.total_adv_completion += adv_stats.completion_tokens
                result.advisor_calls += 1
                if late_rescue:
                    late_rescue_used = True
                if had_error:
                    result.advisor_calls_after_error += 1
                next_tool_hint, next_input_hint = _parse_next_action_from_guidance(guidance)
                result.advisor_guidance.append({
                    "step": step_idx,
                    "after_error": had_error,
                    "trigger": trigger_name,
                    "guidance": guidance,
                    "next_tool_hint": next_tool_hint,
                    "next_input_hint": next_input_hint,
                    "followed_advice": None,  # filled on next step
                })
                step_lat += adv_stats.latency_s
                messages = advisor.integrate_advice(
                    messages,
                    guidance,
                    format_hint=(
                        "Apply this guidance. On your NEXT turn, execute step "
                        "1 of the advisor's NEXT list exactly, unless it is "
                        "clearly impossible. Do NOT repeat any query listed "
                        "in AVOID or any query you already tried. Respond "
                        "with the same JSON schema as before."
                    ),
                )
                # Hedged final should be overridden by new tool calls; clear
                # the 'done' flag to keep the loop going.
                if hedged_final:
                    done = False
                cooldown_until = step_idx + 2
                no_progress_advisor_rounds = 0
            else:
                no_progress_advisor_rounds += 1
                if no_progress_advisor_rounds >= 2:
                    messages.append({
                        "role": "user",
                        "content": (
                            "Stop requesting advisor repeatedly without progress. "
                            "Use remaining evidence to produce your best final answer now."
                        ),
                    })

        # Measure follow-through from prior guidance.
        if (
            result.advisor_guidance
            and result.advisor_guidance[-1].get("followed_advice") is None
            and last_action_dict is not None
            and result.advisor_guidance[-1].get("step") == step_idx - 1
        ):
            hint = result.advisor_guidance[-1].get("next_tool_hint")
            if hint:
                actual_tool = (
                    last_action_dict.get("tool_name", "")
                    if last_action_dict.get("action") == "tool"
                    else None
                )
                followed = actual_tool == hint
                result.advisor_guidance[-1]["followed_advice"] = followed
                result.advisor_first_step_total += 1
                if followed:
                    result.advisor_followed_first_step_count += 1
                else:
                    messages.append({
                        "role": "user",
                        "content": (
                            "You did not follow the advisor's first NEXT step. "
                            "On this turn, execute that first step unless impossible; "
                            "if impossible, explain briefly in thought and choose the closest alternative tool."
                        ),
                    })

        result.step_latencies.append(step_lat)
        if done and not should_escalate:
            result.prediction = answer
            break

        if policy_name.startswith("self_eval") or policy_name.startswith("failure_or_conf_t"):
            # Preserve schema compatibility with downstream analysis.
            result.confidence_scores.append(confidence)

    # If we left the loop with a hedged / empty answer, issue a single
    # force-answer re-prompt before giving up.
    if (
        (result.prediction is None or _is_hedged_answer(result.prediction))
        and not force_answer_used
    ):
        force_answer_used = True
        messages.append({
            "role": "user",
            "content": (
                "You must commit to a final answer now using the evidence "
                "gathered so far. Output ONLY a valid JSON object with "
                "action='final' and a non-empty final_answer that matches "
                "the question's required format. Do not call any tool."
            ),
        })
        t0 = time.perf_counter()
        try:
            resp = client.chat.completions.create(
                model=executor_model,
                messages=messages,
                temperature=temperature,
                seed=seed,
                max_completion_tokens=256,
            )
            latency = time.perf_counter() - t0
            result.total_exec_latency += latency
            usage = resp.usage
            result.total_exec_prompt += usage.prompt_tokens if usage else 0
            result.total_exec_completion += usage.completion_tokens if usage else 0
            forced_text = resp.choices[0].message.content or ""
            forced = _extract_json(forced_text)
            if forced and str(forced.get("action", "")).lower() == "final":
                forced_answer = str(forced.get("final_answer", "")).strip()
                if forced_answer and not _is_hedged_answer(forced_answer):
                    result.prediction = forced_answer
        except Exception:  # noqa: BLE001
            pass

    # Final-answer reformatter pass (cheap) -- coerce to literal GAIA format.
    if result.prediction and not _is_hedged_answer(result.prediction):
        reformatted, pt, ct, lat = _reformat_final_answer(
            client=client,
            model=executor_model,
            question=question,
            candidate=result.prediction,
            temperature=temperature,
            seed=seed,
        )
        result.total_exec_latency += lat
        result.total_exec_prompt += pt
        result.total_exec_completion += ct
        if reformatted:
            result.prediction = reformatted

    result.recovery_success = had_error and post_error_recovered
    return result
